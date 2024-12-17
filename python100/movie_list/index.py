import openpyxl
import json

def load_workbook(file_path):
    return openpyxl.load_workbook(file_path, data_only=False)

def get_hyperlink(cell):
    return cell.hyperlink.target if cell.hyperlink else None

def extract_movie_links(ws):
    movie_links = {}
    for row in ws.iter_rows(min_row=8, values_only=False):
        movie_cell = row[0]
        baidu_cell = row[1]
        kuake_cell = row[2]
        xunlei_cell = row[3]

        movie_name = movie_cell.value

        baidu_link = get_hyperlink(baidu_cell)
        kuake_link = get_hyperlink(kuake_cell)
        xunlei_link = get_hyperlink(xunlei_cell)

        movie_links[movie_name] = {
            '百度网盘': baidu_link,
            '夸克网盘': kuake_link,
            '迅雷云盘': xunlei_link
        }
    return movie_links


# 生成 json
def generate_json(data, filename):
    with open(filename, 'w', encoding='utf-8') as f:
     json.dump(data, f, ensure_ascii=False, indent=2)

# 生成 html
def generate_html(movie_links, output_file):
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write('<html><head><meta charset="utf-8"><title>电影链接</title></head><body>')
        f.write('<h1>电影链接</h1>')
        for movie, links in movie_links.items():
            f.write(f'<h4>{movie}</h4>')
            if links['百度网盘']:
                f.write(f'<a href="{links["百度网盘"]}">百度网盘</a>')
            if links['夸克网盘']:
                f.write(f'<a style="margin-left: 10px" href="{links["夸克网盘"]}">夸克网盘</a>')
            if links['迅雷云盘']:
                f.write(f'<a style="margin-left: 10px" href="{links["迅雷云盘"]}">迅雷云盘</a>')
            f.write('<hr>')
        f.write('</body></html>')

def main():
    file_path = 'movie_list.xlsx'
    output_file = 'movie_list.html'
    wb = load_workbook(file_path)
    ws = wb.active
    movie_links = extract_movie_links(ws)
    generate_json(movie_links, 'movie_list.json')
    generate_html(movie_links, output_file)

if __name__ == "__main__":
    main()