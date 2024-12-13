from openpyxl import load_workbook
from docx import Document
from docx.shared import Inches
import os

# 获取当前工作目录
current_dir = os.path.abspath('employee_data.xlsx')

# 加载Excel文件
workbook = load_workbook('employee_data.xlsx')
sheet = workbook.active

# 加载Word模板
document = Document('offer_letter_template.docx')


for row in sheet.iter_rows(min_row=2): # 从第二行开始读取数据，跳过标题行
  # 获取Excel中的数据
  name = row[0].value
  gender = row[1].value
  position = row[2].value
  start_date = row[3].value.strftime('%Y-%m-%d')

  # 替换Word模板中的占位符
  for paragraph in document.paragraphs:
    paragraph.text = paragraph.text.replace('{姓名}', name)
    paragraph.text = paragraph.text.replace('{职位}', position)
    paragraph.text = paragraph.text.replace('{入职日期}', start_date)
    if gender == '男':
      paragraph.text = paragraph.text.replace('先生/女士', '先生')
    else:
       paragraph.text = paragraph.text.replace('先生/女士', '女士')

  # 保存生成的Word文档
  document.save(f'offer_letter_{name}.docx')

print('录取通知书生成完毕')